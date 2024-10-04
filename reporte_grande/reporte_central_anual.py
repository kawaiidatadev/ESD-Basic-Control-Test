from common.__init__ import *
from reporte_grande.inicializacion import iniciar_reporte
from settings.__init__ import db_path, reporte_c_path


def reporte_grande():
    iniciar_reporte()
    print("Reporte Central ESD")
    generate_report()

def generate_report():
    import sqlite3
    import pandas as pd
    import os
    from openpyxl import load_workbook
    from openpyxl.worksheet.table import Table, TableStyleInfo

    # Conectar a la base de datos SQLite
    conn = sqlite3.connect(db_path)

    # Obtener datos de las tablas y almacenarlos en DataFrames de pandas
    esd_items_df = pd.read_sql_query("SELECT * FROM esd_items", conn)
    personal_esd_df = pd.read_sql_query("SELECT * FROM personal_esd", conn)
    actividades_df = pd.read_sql_query("SELECT * FROM actividades", conn)
    usuarios_elementos_df = pd.read_sql_query("SELECT * FROM usuarios_elementos", conn)
    evidencias_asignacion_df = pd.read_sql_query("SELECT * FROM evidencias_asignacion", conn)
    registro_actividades_df = pd.read_sql_query("SELECT * FROM registro_actividades", conn)
    actividades_registradas_df = pd.read_sql_query("SELECT * FROM actividades_registradas", conn)

    from strings_consultas_db import consulta_elementos_usuarios
    elementos_usuarios_df = pd.read_sql_query(consulta_elementos_usuarios, conn)

    # Cerrar la conexión a la base de datos
    conn.close()

    # Obtener la fecha y hora actual y el nombre de usuario de Windows
    now = datetime.now()
    current_time = now.strftime("%d-%B-%Y %H-%M-%S")
    user_name = getpass.getuser()

    # Crear el nombre del archivo con la fecha y el usuario
    file_name = f'Informe_ESD_{current_time}_{user_name}.xlsx'

    # Ruta completa para guardar el archivo
    save_path = os.path.join(reporte_c_path, file_name)

    # Crear un archivo Excel con múltiples hojas
    with pd.ExcelWriter(save_path, engine='openpyxl') as writer:
        esd_items_df.to_excel(writer, sheet_name='ESD Items', index=False)
        personal_esd_df.to_excel(writer, sheet_name='Personal ESD', index=False)
        actividades_df.to_excel(writer, sheet_name='Actividades', index=False)
        usuarios_elementos_df.to_excel(writer, sheet_name='Usuarios Elementos', index=False)
        evidencias_asignacion_df.to_excel(writer, sheet_name='Evidencias Asignación', index=False)
        registro_actividades_df.to_excel(writer, sheet_name='Registro Actividades', index=False)
        actividades_registradas_df.to_excel(writer, sheet_name='Actividades Registradas', index=False)
        elementos_usuarios_df.to_excel(writer, sheet_name='Usuarios Elementos Todos', index=False)

    # Convertir los datos de cada hoja a tabla
    wb = load_workbook(save_path)

    # Nombres de las hojas
    sheet_names = ['ESD Items', 'Parámetros ESD', 'Personal ESD', 'Actividades',
                   'Usuarios Elementos', 'Evidencias Asignación', 'Registro Actividades',
                   'Actividades Registradas', 'Usuarios Elementos Todos']

    # Recorrer cada hoja para convertir los datos a tabla
    for sheet_name in sheet_names:
        ws = wb[sheet_name]
        # Definir el rango de la tabla (desde A1 hasta la última celda con datos)
        table_ref = f"A1:{chr(65 + ws.max_column - 1)}{ws.max_row}"
        # Crear la tabla
        table = Table(displayName=f"Tabla_{sheet_name.replace(' ', '_')}", ref=table_ref)
        # Establecer un estilo de tabla
        style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                               showLastColumn=False, showRowStripes=True, showColumnStripes=True)
        table.tableStyleInfo = style
        # Agregar la tabla a la hoja
        ws.add_table(table)

    # Guardar el archivo con las tablas creadas
    wb.save(save_path)

    # Notificar la ruta del archivo guardado
    print(f'El archivo ha sido guardado en: {save_path}')
