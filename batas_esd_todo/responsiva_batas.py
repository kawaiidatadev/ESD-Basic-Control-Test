import os
import shutil
import openpyxl
import locale
import win32com.client as win32
from datetime import datetime
import subprocess
import time

# Ruta de la plantilla y la ruta de destino
ruta_plantilla = r"\\mercury\Mtto_Prod\00_Departamento_Mantenimiento\ESD\Software\Recurses\plantilla_responsiva\responsiba.xlsx"
ruta_destino_base = r"\\mercury\Mtto_Prod\00_Departamento_Mantenimiento\ESD\Software\Data\Resposivas generadas"
ruta_bat = r"\\mercury\Mtto_Prod\00_Departamento_Mantenimiento\ESD\Software\Recurses\excel_error\ejecutable.bat"

# Asegúrate de que la configuración regional esté en español (MX)
locale.setlocale(locale.LC_TIME, 'es_MX.UTF-8')


def abrir_excel(archivo_destino):
    """Intenta abrir el archivo de Excel hasta tres veces. Si falla, ejecuta el .bat y lo intenta una vez más."""
    intentos = 0
    while intentos < 3:
        try:
            excel = win32.gencache.EnsureDispatch('Excel.Application')
            excel.Visible = True
            wb = excel.Workbooks.Open(archivo_destino)
            excel.WindowState = win32.constants.xlMaximized
            return True  # Se abrió correctamente
        except Exception as e:
            print(f"Error al intentar abrir Excel: {e}")
            intentos += 1
            time.sleep(1)  # Esperar un segundo antes de volver a intentar
            print(f"Intento {intentos} fallido.")

    # Si después de 3 intentos no se puede abrir, ejecuta el .bat y lo intenta una vez más
    print("Ejecutando el archivo .bat para intentar resolver el error...")
    subprocess.run([ruta_bat], shell=True)

    # Intentar nuevamente abrir Excel
    try:
        excel = win32.gencache.EnsureDispatch('Excel.Application')
        excel.Visible = True
        wb = excel.Workbooks.Open(archivo_destino)
        excel.WindowState = win32.constants.xlMaximized
        return True  # Se abrió correctamente después de ejecutar el .bat
    except Exception as e:
        print(f"Error tras ejecutar el .bat: {e}")
        return False  # No se pudo abrir incluso después del .bat


def generar_responsiba(usuario_id, nombre_usuario, area, linea, tipo_elemento, numero_serie, tamaño):
    print(f"ID del Usuario: {usuario_id}")
    print(f"Nombre Usuario: {nombre_usuario}")
    print(f"Área: {area}")
    print(f"Línea: {linea}")
    print(f"Tipo de Elemento Asignado: {tipo_elemento}")
    print(f"Número de Serie: {numero_serie}")
    print(f"Tamaño: {tamaño}")

    # Obtener la fecha y hora actual en formato español
    fecha_actual = datetime.now()
    fecha_formateada = fecha_actual.strftime("%d-%m-%Y_%H-%M-%S")
    año_actual = fecha_actual.strftime("%Y")

    # Crear la carpeta del año actual si no existe
    ruta_destino = os.path.join(ruta_destino_base, año_actual)
    if not os.path.exists(ruta_destino):
        os.makedirs(ruta_destino)

    # Copiar la plantilla a la nueva ubicación
    archivo_destino = os.path.join(ruta_destino, f"responsiba_{fecha_formateada}.xlsx")
    shutil.copy(ruta_plantilla, archivo_destino)

    # Abrir el archivo copiado
    wb = openpyxl.load_workbook(archivo_destino)
    ws_responsiva = wb['responsiva']  # Seleccionar la hoja 'responsiva'
    ws_datadb = wb['datadb']  # Seleccionar la hoja 'datadb'

    # Leer el número actual en la celda I4 de la hoja 'responsiva'
    numero_actual = ws_responsiva['I4'].value
    try:
        numero_actual = int(numero_actual)
    except (TypeError, ValueError):
        numero_actual = 0  # Valor por defecto si no es un número

    # Incrementar el número por uno
    nuevo_numero = numero_actual + 1

    # Actualizar la celda I4 con el nuevo número en el archivo copiado
    ws_responsiva['I4'] = nuevo_numero

    # Insertar datos en la hoja 'datadb'
    encabezados = [
        ("ID del Usuario", usuario_id),
        ("Nombre Usuario", nombre_usuario),
        ("Área", area),
        ("Línea", linea),
        ("Tipo de Elemento Asignado", tipo_elemento),
        ("Número de Serie", numero_serie),
        ("Tamaño", tamaño)
    ]
    # Agregar el ID del usuario en la celda H2
    ws_datadb['H2'] = usuario_id

    for col_num, (header, value) in enumerate(encabezados, 1):
        ws_datadb.cell(row=1, column=col_num, value=header)
        ws_datadb.cell(row=2, column=col_num, value=value)

    # Guardar el archivo copiado
    wb.save(archivo_destino)
    print(f"Responsiva guardada como {archivo_destino}")

    # Intentar abrir el archivo Excel
    if not abrir_excel(archivo_destino):
        print(
            "No se pudo abrir el archivo Excel después de 3 intentos y de ejecutar el archivo .bat. Operación cancelada.")
        return  # Cancelar operación si no se pudo abrir el archivo Excel

    # Actualizar el número en la plantilla original
    wb_original = openpyxl.load_workbook(ruta_plantilla)
    ws_responsiva_original = wb_original['responsiva']
    ws_responsiva_original['I4'] = nuevo_numero

    wb_original.save(ruta_plantilla)
    print(f"Número en la plantilla original actualizado")
