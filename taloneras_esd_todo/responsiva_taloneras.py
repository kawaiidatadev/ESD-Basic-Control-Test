import os
import shutil
import locale
from datetime import datetime
import openpyxl
import subprocess
import win32com.client as win32
from time import sleep

# Ruta de la plantilla y la ruta de destino
ruta_plantilla = r"\\mercury\Mtto_Prod\00_Departamento_Mantenimiento\ESD\Software\Recurses\plantilla_responsiva\responsiba.xlsx"
ruta_destino_base = r"\\mercury\Mtto_Prod\00_Departamento_Mantenimiento\ESD\Software\Data\Resposivas generadas"
ruta_bat = r"\\mercury\Mtto_Prod\00_Departamento_Mantenimiento\ESD\Software\Recurses\excel_error\ejecutable.bat"

# Asegúrate de que la configuración regional esté en español (MX)
locale.setlocale(locale.LC_TIME, 'es_MX.UTF-8')


def abrir_excel_maximizado(archivo):
    """
    Intenta abrir un archivo Excel en modo maximizado.
    Si no se puede abrir después de 3 intentos, ejecuta un .bat y vuelve a intentar.
    Si sigue fallando, se cancela la operación.
    """
    intentos = 0
    max_intentos = 3
    while intentos < max_intentos:
        try:
            excel = win32.gencache.EnsureDispatch('Excel.Application')
            wb = excel.Workbooks.Open(archivo)
            excel.Visible = True
            excel.WindowState = win32.constants.xlMaximized
            excel.Windows(1).Activate()
            print("Archivo Excel abierto y maximizado correctamente.")
            return True  # Éxito
        except Exception as e:
            print(f"Error al abrir Excel en el intento {intentos + 1}: {e}")
            intentos += 1
            sleep(1)  # Espera un segundo antes de reintentar

    # Ejecutar el archivo .bat si después de 3 intentos no se abrió Excel
    print("No se pudo abrir Excel después de 3 intentos. Ejecutando archivo .bat...")
    try:
        subprocess.call([ruta_bat], shell=True)
        print(".bat ejecutado, intentando abrir Excel nuevamente.")
        try:
            excel = win32.gencache.EnsureDispatch('Excel.Application')
            wb = excel.Workbooks.Open(archivo)
            excel.Visible = True
            excel.WindowState = win32.constants.xlMaximized
            excel.Windows(1).Activate()
            print("Archivo Excel abierto y maximizado correctamente.")
            return True  # Éxito después de ejecutar el .bat
        except Exception as e:
            print(f"Error al abrir Excel después de ejecutar el .bat: {e}")
            return False  # Fallo final
    except Exception as bat_error:
        print(f"Error al ejecutar el .bat: {bat_error}")
        return False  # Fallo final


def generar_responsiva_taloneras(usuario_id, nombre_usuario, area, linea, tipo_elemento, numero_serie='N-A',
                                 tamaño='N-A'):
    """
    Genera la responsiva taloneras y actualiza la plantilla. Si no se puede abrir Excel, no se registran cambios.
    """
    print(f"ID del Usuario: {usuario_id}")
    print(f"Nombre Usuario: {nombre_usuario}")
    print(f"Área: {area}")
    print(f"Línea: {linea}")
    print(f"Tipo de Elemento Asignado: {tipo_elemento}")
    print(f"Número de Serie: {numero_serie}")
    print(f"Tamaño: {tamaño}")

    # Obtener la fecha y hora actual en formato español
    fecha_actual = datetime.now()
    fecha_formateada = fecha_actual.strftime("%d-%m-%Y_%H-%M-%S")
    año_actual = fecha_actual.strftime("%Y")

    # Crear la carpeta del año actual si no existe
    ruta_destino = os.path.join(ruta_destino_base, año_actual)
    if not os.path.exists(ruta_destino):
        os.makedirs(ruta_destino)

    # Copiar la plantilla a la nueva ubicación
    archivo_destino = os.path.join(ruta_destino, f"responsiba_{fecha_formateada}.xlsx")
    shutil.copy(ruta_plantilla, archivo_destino)

    try:
        # Abrir el archivo copiado
        wb = openpyxl.load_workbook(archivo_destino)
        ws_responsiva = wb['responsiva']  # Seleccionar la hoja 'responsiva'
        ws_datadb = wb['datadb']  # Seleccionar la hoja 'datadb'

        # Leer el número actual en la celda I4 de la hoja 'responsiva'
        numero_actual = ws_responsiva['I4'].value
        try:
            numero_actual = int(numero_actual)
        except (TypeError, ValueError):
            numero_actual = 0  # Valor por defecto si no es un número

        # Incrementar el número por uno
        nuevo_numero = numero_actual + 1

        # Actualizar la celda I4 con el nuevo número en el archivo copiado
        ws_responsiva['I4'] = nuevo_numero

        # Insertar datos en la hoja 'datadb'
        encabezados = [
            ("ID del Usuario", usuario_id),
            ("Nombre Usuario", nombre_usuario),
            ("Área", area),
            ("Línea", linea),
            ("Tipo de Elemento Asignado", tipo_elemento),
            ("Número de Serie", numero_serie),
            ("Tamaño", tamaño)
        ]

        for col_num, (header, value) in enumerate(encabezados, 1):
            ws_datadb.cell(row=1, column=col_num, value=header)
            ws_datadb.cell(row=2, column=col_num, value=value)

        # Guardar el archivo copiado
        wb.save(archivo_destino)
        print(f"Responsiva guardada como {archivo_destino}")

        # Intentar abrir el archivo automáticamente con ventana maximizada
        if not abrir_excel_maximizado(archivo_destino):
            print("Operación cancelada debido a problemas al abrir el archivo Excel.")
            return  # Cancela toda la operación si no se pudo abrir Excel

        # Actualizar el número en la plantilla original
        wb_original = openpyxl.load_workbook(ruta_plantilla)
        ws_responsiva_original = wb_original['responsiva']
        ws_responsiva_original['I4'] = nuevo_numero
        wb_original.save(ruta_plantilla)
        print(f"Número en la plantilla original actualizado")

    except Exception as e:
        print(f"Error al procesar el archivo Excel: {e}")
